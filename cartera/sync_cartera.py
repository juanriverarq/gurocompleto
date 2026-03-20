#!/usr/bin/env python3
"""
Sync cartera_items from SoftSeguros Excel exports.
Reads 4 Excel files, matches by IDENTIFICADOR = softseguros_pago_id,
and generates SQL UPDATE statements to fill missing fields.

Usage:
  python3 sync_cartera.py          # dry-run (prints stats)
  python3 sync_cartera.py --execute # generates SQL file and runs it
"""

import sys
import os
import openpyxl
from datetime import datetime

EXCEL_DIR = os.path.dirname(os.path.abspath(__file__))
SQL_OUTPUT = os.path.join(EXCEL_DIR, 'sync_cartera_updates.sql')

# Excel column → DB column
COL_MAP = {
    'IDENTIFICADOR': 'softseguros_pago_id',
    'NÚMERO ANEXO': 'anexo_numero',
    'NÚMERO REMISIÓN': 'numero_remision',
    'CATEGORÍAS PÓLIZA': 'categorias_poliza',
    'CATEGORÍAS CLIENTE': 'categorias_cliente',
    'VALOR PRIMA NETA': 'prima_neta',
    'VALOR NETO A PAGAR': 'valor_neto_a_pagar',
    'PRIMA TOTAL DEL PAGO': 'prima_total_pago',
    'VALOR PRIMA TOTAL': 'prima_total',
    'SALDO PENDIENTE OFICINA': 'saldo_pendiente_oficina',
    'SALDO PENDIENTE ASEGURADORA': 'saldo_pendiente_aseguradora',
    'COMISIÓN A RECIBIR': 'comision_a_recibir',
    'COMISIÓN VENDEDOR': 'comision_vendedor',
    'COMISIÓN RECIBIDA': 'comision_recibida',
    'COMISIÓN PAGADA VENDEDOR': 'comision_pagada_vendedor',
    'VALOR RECAUDADO EN OFICINA': 'valor_recaudado_oficina',
    'VALOR PAGADO EN ASEGURADORA': 'valor_pagado_aseguradora',
    'PORCENTAJE DE COMISIÓN': 'porcentaje_comision',
    'PORCENTAJE DE COMISIÓN ANEXO': 'porcentaje_comision_anexo',
    'CÓDIGO RADICACIÓN': 'codigo_radicacion',
    'FECHA LÍMITE DE PAGO': 'fecha_limite_pago',
    'FECHA COMPROMISO DE PAGO': 'fecha_compromiso_pago',
    'FECHA RECAUDO EN OFICINA': 'fecha_recaudado_oficina',
    'FECHA RECAUDADO EN OFICINA': 'fecha_recaudado_oficina',
    'FECHA REALIZÓ PAGO EN ASEGURADORA': 'fecha_pago_aseguradora',
    'FECHA COMISIONADA': 'fecha_comisionada',
    'FECHA PAGADA VENDEDOR': 'fecha_pagada_vendedor',
    'USUARIO COMISIONO': 'usuario_comisiono',
    'MEDIO DE PAGO': 'medio_pago',
    'CÓDIGO CONTABLE': 'codigo_contable',
    'TIPO MONEDA': 'moneda',
    'FECHA INICIO VIGENCIA PÓLIZA': 'fecha_inicio_vigencia',
    'FECHA FIN VIGENCIA PÓLIZA': 'fecha_fin_vigencia',
    'RECAUDADO ASEGURADORA PENDIENTE POR COBRAR AL CLIENTE': 'recaudado_aseg_pendiente_cliente',
    'ÚLTIMA OBSERVACIÓN BITACORA': 'observacion_bitacora',
    'OBSERVACIONES PAGO': 'observaciones_pago',
}

NUMERIC_COLS = {
    'prima_neta', 'valor_neto_a_pagar', 'prima_total_pago', 'prima_total',
    'saldo_pendiente_oficina', 'saldo_pendiente_aseguradora',
    'comision_a_recibir', 'comision_vendedor', 'comision_recibida',
    'comision_pagada_vendedor', 'valor_recaudado_oficina', 'valor_pagado_aseguradora',
    'porcentaje_comision', 'porcentaje_comision_anexo',
}

DATE_COLS = {
    'fecha_limite_pago', 'fecha_compromiso_pago', 'fecha_recaudado_oficina',
    'fecha_pago_aseguradora', 'fecha_comisionada', 'fecha_pagada_vendedor',
    'fecha_inicio_vigencia', 'fecha_fin_vigencia',
}

FLAGS_BY_ESTADO = {
    'por_cobrar': (0, 0, 0),
    'por_pagar': (1, 0, 0),
    'comision_por_cobrar': (1, 1, 0),
    'comision_recibida': (1, 1, 1),
}

FILES = [
    {'pattern': 'por_cobrar', 'estado': 'por_cobrar'},
    {'pattern': 'por_pagar', 'estado': 'por_pagar'},
    {'pattern': 'comisiones_por_cobrar', 'estado': 'comision_por_cobrar'},
    {'pattern': 'comisiones_recibidas', 'estado': 'comision_recibida'},
]


def find_file(pattern):
    for f in os.listdir(EXCEL_DIR):
        if f.startswith(pattern) and f.endswith('.xlsx') and not f.startswith('~'):
            return os.path.join(EXCEL_DIR, f)
    return None


def escape_sql(val):
    if val is None:
        return 'NULL'
    s = str(val).replace("\\", "\\\\").replace("'", "\\'")
    return f"'{s}'"


def format_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(val, str) and val.strip():
        try:
            return datetime.fromisoformat(val.strip()).strftime('%Y-%m-%d %H:%M:%S')
        except:
            return None
    return None


def process_file(filepath, estado, sql_file=None):
    print(f"\n📄 Processing: {os.path.basename(filepath)} → estado={estado}")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    # Read headers from first row
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h).strip() if h else '' for h in header_row]

    # Map column index → db column
    col_indices = {}
    for i, h in enumerate(headers):
        if h in COL_MAP:
            col_indices[i] = COL_MAP[h]

    flags = FLAGS_BY_ESTADO[estado]
    count = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Get softseguros_pago_id
        ss_id = None
        data = {}
        for i, db_col in col_indices.items():
            val = row[i] if i < len(row) else None
            if db_col == 'softseguros_pago_id':
                ss_id = val
                continue
            data[db_col] = val

        if not ss_id:
            skipped += 1
            continue

        # Build SET clause
        sets = []

        for col, val in data.items():
            if val is None or (isinstance(val, str) and val.strip() == ''):
                continue

            if col in NUMERIC_COLS:
                try:
                    fval = float(val)
                    sets.append(f"`{col}` = {fval}")
                except (ValueError, TypeError):
                    continue
            elif col in DATE_COLS:
                dval = format_date(val)
                if dval:
                    sets.append(f"`{col}` = '{dval}'")
            elif col == 'recaudado_aseg_pendiente_cliente':
                bval = 1 if str(val).strip().lower() in ('si', 'sí', '1', 'true') else 0
                sets.append(f"`{col}` = {bval}")
            elif col == 'dias_vencidos':
                try:
                    sets.append(f"`{col}` = {int(val)}")
                except:
                    continue
            else:
                sets.append(f"`{col}` = {escape_sql(val)}")

        # Always set estado + flags
        sets.append(f"`estado_cartera` = '{estado}'")
        sets.append(f"`recaudado_en_oficina` = {flags[0]}")
        sets.append(f"`recaudado_aseguradora` = {flags[1]}")
        sets.append(f"`comisionada` = {flags[2]}")

        if sets:
            sql = f"UPDATE `cartera_items` SET {', '.join(sets)} WHERE `softseguros_pago_id` = {int(ss_id)};\n"
            if sql_file:
                sql_file.write(sql)
            count += 1

        if count % 5000 == 0 and count > 0:
            print(f"  Progress: {count} rows...")

    wb.close()
    print(f"  ✅ {count} updates generated, {skipped} skipped")
    return count


def main():
    execute = '--execute' in sys.argv
    total = 0

    if execute:
        print(f"🚀 EXECUTE MODE — generating SQL file: {SQL_OUTPUT}")
        sql_file = open(SQL_OUTPUT, 'w', encoding='utf-8')
        sql_file.write("SET NAMES utf8mb4;\n")
        sql_file.write("SET FOREIGN_KEY_CHECKS = 0;\n\n")
    else:
        print("🔍 DRY-RUN MODE — counting matches only")
        sql_file = None

    for fc in FILES:
        filepath = find_file(fc['pattern'])
        if not filepath:
            print(f"  ⚠️ No file for pattern: {fc['pattern']}")
            continue
        total += process_file(filepath, fc['estado'], sql_file)

    if execute:
        sql_file.write("\nSET FOREIGN_KEY_CHECKS = 1;\n")
        sql_file.close()
        print(f"\n📝 SQL file written: {SQL_OUTPUT} ({total} statements)")
        print(f"   Size: {os.path.getsize(SQL_OUTPUT) / 1024 / 1024:.1f} MB")
    else:
        print(f"\n📊 Total rows that would be updated: {total}")
        print("   Add --execute to generate the SQL file")


if __name__ == '__main__':
    main()
